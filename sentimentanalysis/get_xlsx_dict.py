from openpyxl import load_workbook
def get_dict(xlsx):
    pos_dict = {}
    neg_dict = {}
    workbook = load_workbook(xlsx)
    actsheet = workbook['Sheet1']
    row_num = 0
    for row in actsheet.iter_rows():
        # if row_num > 99:
        #     break
        if row_num > 0:
            if row[4].value[0] == u'P':
                pos_dict[row[0].value] = row[5].value
                #print row[0].value, row[6].value
            elif row[4].value[0] == u'N':
                neg_dict[row[0].value] = row[5].value
                #print row[0].value, row[6].value
        row_num += 1
    return {"pos": pos_dict, "neg": neg_dict}


#get_dict("dict.xlsx")