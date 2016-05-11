#! /usr/bin/env python
# coding: utf-8
from spider.sns import get_content
from openpyxl import Workbook
workbook = Workbook()
worksheet = workbook.active
cells = worksheet["A1":"E100"]
print worksheet.rows
for row in range(1, 101):
    worksheet.cell(row=row, column=1, value=row)
    worksheet.cell(row=row, column=2, value=get_content.contents[row - 1])
    worksheet.cell(row=row, column=3, value=str(get_content.human[row - 1]))
    worksheet.cell(row=row, column=4, value=str(get_content.robot[row - 1]))
    worksheet.cell(row=row, column=5, value=str(get_content.correct[row - 1]).decode('utf-8'))
workbook.save("sta.xlsx")